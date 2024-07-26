from flask import Flask, render_template, jsonify, send_from_directory
from threading import Thread
import cv2
import pickle
import numpy as np
import matplotlib.pyplot as plt
from openpyxl import load_workbook, Workbook
import os
import datetime

# Set Matplotlib backend to Agg for non-GUI environments
import matplotlib
matplotlib.use('Agg')

app = Flask(__name__)

url = 'http://192.168.55.139:8080/video'
cap = cv2.VideoCapture(0)

with open('CarParkPos', 'rb') as f:
    posList = pickle.load(f)

width, height = 290, 180
prev_free_spots = 0

def checkParkingSpace(imgPro):
    spaceCounter = 0
    occupiedSlots = []

    for i, pos in enumerate(posList):
        if len(pos) != 3:
            print(f'Invalid pos: {pos}')  # Debugging statement
            continue

        x, y, _ = pos

        imgCrop = imgPro[y:y + height, x:x + width]
        count = cv2.countNonZero(imgCrop)

        if count < 900:
            spaceCounter += 1
        else:
            occupiedSlots.append(chr(ord("A") + i))

    emptySlots = [chr(ord("A") + i) for i in range(len(posList)) if chr(ord("A") + i) not in occupiedSlots]

    return spaceCounter, emptySlots

def update_excel(free_spots, empty_spots):
    filename = 'parking_data.xlsx'
    file_exists = os.path.isfile(filename)

    wb = Workbook() if not file_exists else load_workbook(filename)
    ws = wb.active
    if not file_exists:
        ws.append(['Date', 'Free Spots'])

    ws.append([datetime.datetime.now(), free_spots])
    wb.save(filename)

def generate_graph():
    filename = 'parking_data.xlsx'
    try:
        if not os.path.isfile(filename):
            print("Excel file does not exist.")
            return
        
        wb = load_workbook(filename)
        ws = wb.active

        dates = []
        free_spots = []

        for row in ws.iter_rows(min_row=2, values_only=True):
            dates.append(row[0])
            free_spots.append(row[1])

        plt.figure(figsize=(15, 10))  # Increase figure size
        plt.plot(dates, free_spots, marker='o', linestyle='-', color='b')
        plt.title('Parking Space Trend')
        plt.xlabel('Date')
        plt.ylabel('Free Spots')
        plt.grid(True)
        plt.gcf().autofmt_xdate()  # Format date labels
        plt.tight_layout()

        if not os.path.exists('static'):
            os.makedirs('static')

        plt.savefig('static/parking_trend.png')
        plt.close()
    except Exception as e:
        print(f"Error generating graph: {e}")


def capture_video():
    global prev_free_spots
    cv2.namedWindow("Parking Detection")  # Create a window with a specific name
    cv2.moveWindow("Parking Detection", 100, 100)  # Move the window to a visible location

    while True:
        success, img = cap.read()
        if not success:
            print("Error capturing video frame.")
            break

        imgGray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
        imgBlur = cv2.GaussianBlur(imgGray, (3, 3), 1)
        imgThreshold = cv2.adaptiveThreshold(imgBlur, 255, cv2.ADAPTIVE_THRESH_GAUSSIAN_C,
                                             cv2.THRESH_BINARY_INV, 25, 16)
        imgMedian = cv2.medianBlur(imgThreshold, 5)
        kernel = np.ones((3, 3), np.uint8)
        imgDilate = cv2.dilate(imgMedian, kernel, iterations=1)

        free_spots, empty_spots = checkParkingSpace(imgDilate)

        cv2.putText(img, f'Free Spots: {free_spots}', (10, 30), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 255, 0), 2)
        cv2.putText(img, f'Empty Spots: {", ".join(empty_spots)}', (10, 70), cv2.FONT_HERSHEY_SIMPLEX, 1, (0, 0, 255), 2)

        for i, pos in enumerate(posList):
            if len(pos) != 3:
                print(f'Invalid pos for rectangle: {pos}')  # Debugging statement
                continue

            x, y, _ = pos
            color = (0, 255, 0) if chr(ord("A") + i) in empty_spots else (0, 0, 255)
            thickness = 5 if chr(ord("A") + i) in empty_spots else 2
            cv2.rectangle(img, (x, y), (x + width, y + height), color, thickness)

        cv2.imshow("Parking Detection", img)

        if free_spots != prev_free_spots:
            print(f'Free Spots: {free_spots}, Empty Spots: {", ".join(empty_spots)}')

            with open('parking_info.txt', 'a') as file:
                file.write(f'Free Spots: {free_spots}, Empty Spots: {", ".join(empty_spots)}\n')

            update_excel(free_spots, empty_spots)
            generate_graph()

        prev_free_spots = free_spots

        if cv2.waitKey(1) & 0xFF == ord('q'):
            break

    cap.release()
    cv2.destroyAllWindows()

@app.route('/prototype')
def home():
    return render_template('prototype.html')

@app.route('/')
def index():
    return render_template('index.html')

@app.route('/get_overall_status')
def get_overall_status():
    try:
        with open('parking_info.txt', 'r') as file:
            lines = file.readlines()
            latest_info = lines[-1]  # Get the latest information
            free_spots = int(latest_info.split(':')[1].split(',')[0].strip())
            empty_spots = latest_info.split(':')[2].strip().split(', ')
            return jsonify({'free': free_spots, 'total': len(posList), 'empty': empty_spots})
    except Exception as e:
        print(f"Error reading parking_info.txt: {e}")
        return jsonify({'free': prev_free_spots, 'total': len(posList), 'empty': []})

@app.route('/get_graph')
def get_graph():
    return send_from_directory('static', 'parking_trend.png')

if __name__ == '__main__':
    video_thread = Thread(target=capture_video)
    video_thread.start()
    app.run(debug=True)
